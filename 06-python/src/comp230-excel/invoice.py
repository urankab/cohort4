import openpyxl

# Load data from worksheet
# Use dictionaries to store customer, invoices, lines, product
# Ask user for invoice to print
# Allow for (don't code): total invoiced to each client
# invoiced amount each day, invoices by client

invoice_list = {}
product_list = []
customer_list = []
invoice_line_list = []

def sheets_to_dict():
    path = r'C:\code\cohort4\06-python\src\comp230-excel\comp_230.xlsx'
    wb = openpyxl.load_workbook(path)
    worksheet_names = wb.sheetnames
    sheet_index = worksheet_names.index('Invoices')
    wb.active = sheet_index

    for row in wb.active.iter_rows(min_row=2, values_only=True):
        # invoices = {
        #     'invoice_no': row[0],
        #     'customer_id': row[1],
        #     'invoice_date': row[2],
        #     'payment_method': row[3],
        #     'total_price': row[4]
        # }
        # invoice_list.append(invoices)

        # Play ~

        #  ~

    # sheet_index = worksheet_names.index('Products')
    # wb.active = sheet_index

    # for row in wb.active.iter_rows(min_row=2, values_only=True):
    #     products = {
    #         'product_id': row[0],
    #         'name': row[1],
    #         'description': row[2],
    #         'quantity': row[3],
    #         'cost': row[4]
    #     }
    #     product_list.append(products)

    # sheet_index = worksheet_names.index('Customers')
    # wb.active = sheet_index

    # for row in wb.active.iter_rows(min_row=2, values_only=True):
    #     customer = {
    #         'customer_id': row[0],
    #         'f_name': row[1],
    #         'l_name': row[2],
    #         'phone': row[3],
    #         'email': row[4],
    #         'street': row[5],
    #         'city': row[6],
    #         'zipcode': row[7]
    #     }
    #     customer_list.append(customer)
    
    # sheet_index = worksheet_names.index('Invoice Line Items')
    # wb.active = sheet_index

    # for row in wb.active.iter_rows(min_row=2, values_only=True):
    #     inv_line = {
    #         'invoice_no': row[0],
    #         'product_id': row[1],
    #         'quantity': row[2],
    #         'price': row[3]
    #     }
    #     invoice_line_list.append(inv_line)

    # print(invoice_list)
    # print(product_list)
    # print(customer_list)
    # print(invoice_line_list)

def invoice_generator():
    sheets_to_dict()
    # user_input = input('Please enter the invoice number: ')
    print(invoice_list)

invoice_generator()