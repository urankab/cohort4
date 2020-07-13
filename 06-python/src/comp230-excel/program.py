import functions
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook


def requestInvoice():
    functions.sheets_to_dict()

    title_style = Font(name='Arial Rounded MT Bold', sz=25, b=True, color='A872DD')
    header_style = Font(name='Arial Rounded MT Bold', sz=14, b=True, color='8E44AD')
    customer_style = Font(name='Arial Rounded MT Bold', sz=12, b=True, color='5A49A5')
    tag_style = Font(name='Arial Rounded MT Bold', sz=14, b=True, color='8E44AD')
    item_style = Font(name='Arial Rounded MT Bold', sz=10, b=True, color='5A49A5')
    total_style = Font(name='Arial Rounded MT Bold', sz=14, b=True, color='A872DD')

    inv_no = int(input('Please enter an invoice number: '))

    try:
        if inv_no == functions.invoice_list[inv_no]['invoice_no']:
            # Customer Information
            customer_id = functions.invoice_list[inv_no]['customer_id']
            customer_name = str(functions.customer_list[customer_id]['f_name'] + ' ' +
                                functions.customer_list[customer_id]['l_name'])
            customer_phone = str(functions.customer_list[customer_id]['phone'])
            customer_email = str(functions.customer_list[customer_id]['email'])
            customer_residence = str(functions.customer_list[customer_id]['street'] + ', ' +
                                    functions.customer_list[customer_id]['city'])
            customer_zip = str(functions.customer_list[customer_id]['zipcode'])

            # Invoice Lines & Products
            invoice_lines = functions.invoice_line_list
            product_ids = []
            quantities = []

            for key in invoice_lines:
                if inv_no == invoice_lines[key]['invoice_no']:
                    product_ids.append(invoice_lines[key]['product_id'])
                    quantities.append(invoice_lines[key]['quantity'])

            product_names = []
            product_descs = []
            product_costs = []
            amounts = []

            if len(product_ids) > 1:
                for index in range(len(product_ids)):
                    ids = product_ids[index]
                    product_names.append(functions.product_list[ids]['name'])
                    product_descs.append(functions.product_list[ids]['description'])
                    product_costs.append(functions.product_list[ids]['cost'])

            # Amount ( quantity * unit price )
            for i in range(0, len(product_costs)):
                amounts.append(product_costs[i] * quantities[i])

            # ----- WRITING INVOICE ----------------------------------------------

            workbook = Workbook()
            sheet = workbook.active

            # CUSTOMER INFO

            sheet['E2'] = "Uranka's Outdoors - Invoice"
            sheet['E2'].alignment = Alignment(horizontal='center')
            sheet['E2'].font = title_style

            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 20

            for i in range(1, 8):
                sheet['A' + str(i)].font = header_style
                sheet['A' + str(i)].alignment = Alignment(horizontal='center')
                
            sheet['A1'] = 'Invoice No.'
            sheet['A2'] = 'Date: '
            sheet['A3'] = 'Customer: '
            sheet['A4'] = 'Contact: '
            sheet['A6'] = 'Residence: '
            sheet['A7'] = 'Zipcode: '

            for i in range(1, 8):
                sheet['B' + str(i)].font = customer_style
                sheet['B' + str(i)].alignment = Alignment(horizontal='left')

            sheet['B1'] = functions.invoice_list[inv_no]['invoice_no']
            sheet['B2'] = functions.invoice_list[inv_no]['invoice_date']
            sheet['B3'] = customer_name
            sheet['B4'] = customer_phone
            sheet['B5'] = customer_email
            sheet['B6'] = customer_residence
            sheet['B7'] = customer_zip

            # INVOICE LINES

            # Center Titles for columns
            for col in sheet.iter_cols(min_row=11,max_col=6, max_row=11):
                for cell in col:
                    cell.font = tag_style
                    cell.alignment = Alignment(horizontal='center')
                    
            # Center item lines
            for col in sheet.iter_cols(min_row=12,max_col=5, max_row=12+len(product_names)):
                for cell in col:
                    cell.font = item_style
                    cell.alignment = Alignment(horizontal='center')


            sheet['A11'] = 'Item'
            for i in range(0, len(product_names)):
                sheet.cell(row = i + 12, column = 1).value = product_names[i]
                sheet.cell(row = i + 12, column = 1).alignment = Alignment(horizontal='center')


            sheet['B11'] = 'Description'
            for i in range(0, len(product_descs)):
                sheet.cell(row=i + 12, column = 2).value = product_descs[i]
                sheet.cell(row=i + 12, column = 2).alignment = Alignment(horizontal='center')


            sheet['C11'] = 'Quantity'
            for i in range(0, len(quantities)):
                sheet.cell(row=i + 12, column = 3).value = quantities[i]
                sheet.cell(row=i + 12, column = 3).alignment = Alignment(horizontal='center')


            sheet['D11'] = 'Unit Price'
            for i in range(0, len(product_costs)):
                sheet.cell(row=i + 12, column = 4).value = f'${product_costs[i]}'
                sheet.cell(row=i + 12, column = 4).alignment = Alignment(horizontal='center')


            sheet['E11'] = 'Amount'
            for i in range(0, len(amounts)):
                sheet.cell(row=i + 12, column = 5).value = f'${amounts[i]}'
                sheet.cell(row=i + 12, column = 5).alignment = Alignment(horizontal='center')

            sheet['F11'] = 'Amount Due'
            sheet['F' + str(12 + len(amounts))].value = f'${sum(amounts)}'
            sheet['F'+str(12+len(amounts))].alignment = Alignment(horizontal='center')
            sheet['F'+str(12+len(amounts))].font = total_style


            workbook.save(
                f'/code/cohort4/06-python/src/comp230-excel/invoice_{inv_no}.xlsx')
            print(f'Invoice has been written to invoice_{inv_no}.xlsx')

    except:
        print(f'There is no such invoice number')
        return requestInvoice()


requestInvoice()
