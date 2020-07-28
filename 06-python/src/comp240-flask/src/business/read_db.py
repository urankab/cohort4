import openpyxl

invoice_list = {}
product_list = {}
customer_list = {}
invoice_line_list = {}

def sheets_to_dict():
    path = r"C:\code\cohort4\06-python\src\comp230-excel\comp_230.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb["Invoices"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        invoices = {
            "invoice_no": row[0],
            "customer_id": row[1],
            "invoice_date": row[2],
            "payment_method": row[3],
            "total_price": row[4]
        }
        invoice_list[row[0]] = invoices    
    # print(invoice_list)
    # return invoice_list


    sheet = wb["Products"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        products = {
            "product_id": row[0],
            "name": row[1],
            "description": row[2],
            "quantity": row[3],
            "cost": row[4]
        }
        product_list[row[0]] = products
    # print(product_list)
    # return product_list


    sheet = wb["Customers"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        customer = {
            "customer_id": row[0],
            "f_name": row[1],
            "l_name": row[2],
            "phone": row[3],
            "email": row[4],
            "street": row[5],
            "city": row[6],
            "zipcode": row[7]
        }
        customer_list[row[0]] = customer
    # print(customer_list)
    # return customer_list

    sheet = wb["Line Items"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        inv_line = {
            "line_id": row[0],
            "invoice_no": row[1],
            "product_id": row[2],
            "quantity": row[3],
            "price": row[4]
        }
        invoice_line_list[row[0]] = inv_line

    return customer_list, invoice_list, invoice_line_list, product_list


def toTestCustomerValues():
    sheets_to_dict()
    return customer_list[1]["f_name"]

def toTestInvoiceValues():
    sheets_to_dict()
    return invoice_list[1]

def toTestInvoiceLineValues():
    sheets_to_dict()
    return invoice_line_list[1]

def toTestProductValues():
    sheets_to_dict()
    return product_list[100]

toTestCustomerValues()
toTestInvoiceValues()
toTestInvoiceLineValues()
toTestProductValues()